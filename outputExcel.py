import openpyxl
from DPSParser import *
from sklearnCluster import *

wb = openpyxl.load_workbook('similarity_copy1.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#sheet.title = 'All Similarity'
dpsc = DPSParser()
#first500case = dpsc[0:500]
myDict = {}
myclassDict = {}
dictify(dpsc, myDict, myclassDict)

myclassInfo = []
myClassList = []

dicttolist(myclassDict, myclassInfo, myClassList)

similarity_matrix = get_similarity_matrix(myclassInfo)
#print(type(myclassInfo[1]))
#print(len(myclassInfo[1]))
#print(get_similarity(myclassInfo[0], myclassInfo[1]))

"""
trafficcase = []
for case in dpsc:
    if case.ccClass == "TRAFFIC":
        trafficcase.append(case)
for i in range(50):
    for x in range(50):
        sheet.cell(row=i+2,column=x+2).value = get_similarity(dpsc[i], dpsc[x])
    sheet.cell(row=1, column=i+2).value = str(dpsc[i].casenbr)+ " :" +str(dpsc[i].ccDescription)
    sheet.cell(row=i+2, column=1).value = str(dpsc[i].casenbr)+ " :" +str(dpsc[i].ccDescription)"""

for i in range(len(myclassInfo)):
    for x in range(len(myclassInfo)):
        sheet.cell(row=i+2,column=x+2).value = similarity_matrix[i][x]
    sheet.cell(row=1, column=i+2).value = myClassList[i]
    sheet.cell(row=i+2, column=1).value = myClassList[i]

wb.save('similarity_copy1.xlsx')