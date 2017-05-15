import openpyxl
from DPSParser import *
from sklearnCluster import *

wb = openpyxl.load_workbook('similarity_copy1.xlsx')
sheet = wb.get_sheet_by_name('Sheet3')
#sheet.title = 'All Similarity'
dpsc = DPSParser()
trafficcase = []
"""for case in dpsc:
    if case.ccClass == "TRAFFIC":
        trafficcase.append(case)"""
for i in range(50):
    for x in range(50):
        sheet.cell(row=i+2,column=x+2).value = get_similarity(dpsc[i], dpsc[x])
    sheet.cell(row=1, column=i+2).value = str(dpsc[i].casenbr)+ " :" +str(dpsc[i].ccDescription)
    sheet.cell(row=i+2, column=1).value = str(dpsc[i].casenbr)+ " :" +str(dpsc[i].ccDescription)
wb.save('similarity_copy1.xlsx')